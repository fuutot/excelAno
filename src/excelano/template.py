from __future__ import annotations

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from excelano.schema import Schema, SchemaValidationError


class Template(pd.DataFrame):
    """アノテーション用のエクセルファイルのテンプレートを作成するためのクラス"""

    _metadata = ["annotation_cols", "id_cols", "schema"]  # pandasのDataFrameに属性を保持させるための変数

    @property
    def _constructor(self):
        return Template

    @staticmethod
    def from_dataframe(
        df: pd.DataFrame, id_cols: list[str], annotation_cols: list[str], schema: Schema | None = None
    ) -> Template:
        """
        DataFrameからテンプレートを作成するメソッド
        args:
            df: テンプレートの元となるDataFrame
            id_cols: 評価対象を一意に識別する列名のリスト
            annotation_cols: アノテーション対象列名のリスト．
            schema: バリデーション用のSchemaオブジェクト（任意）
        returns:
            Template: 作成したテンプレート
        """
        template = Template(df)

        # id_colsで一意に識別できるか確認
        if template.duplicated(subset=id_cols).any():
            raise ValueError("id_colsで指定された列の組み合わせで一意に識別できません。")

        # アノテーション対象の列に値が入っていないか確認
        if template[annotation_cols].notnull().any().any():
            raise ValueError(
                "annotation_colsで指定された列に値が入っています。アノテーション対象の列は空にしてください。"
            )

        # Schemaによるバリデーション
        if schema is not None:
            errors = schema.validate(df)
            if errors:
                raise SchemaValidationError(errors)

        template.annotation_cols = annotation_cols
        template.id_cols = id_cols
        template.schema = schema
        return template

    @staticmethod
    def from_csv(
        file_path: str, id_cols: list[str], annotation_cols: list[str], schema: Schema | None = None, **kwargs
    ) -> Template:
        """
        CSVファイルからテンプレートを作成するメソッド
        args:
            file_path: CSVファイルのパス
            id_cols: 評価対象を一意に識別する列名のリスト
            annotation_cols: アノテーション対象列名のリスト．
            schema: バリデーション用のSchemaオブジェクト（任意）
            **kwargs: pandas.read_csvに渡す引数
        returns:
            Template: 作成したテンプレート
        """
        df = pd.read_csv(file_path, **kwargs)
        return Template.from_dataframe(df, id_cols=id_cols, annotation_cols=annotation_cols, schema=schema)

    def to_excel(self, file_path: str) -> None:
        """
        テンプレートをエクセルファイルに保存するメソッド
        args:
            file_path: 保存するエクセルファイルのパス
        """
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # DataFrameをエクセルファイルに書き込む
            super().to_excel(writer, index=False)

            worksheet: Worksheet = writer.sheets["Sheet1"]

            # 罫線のスタイルを定義
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # ヘッダーの書式を設定
            bold_font = Font(bold=True, color="FFFFFF")  # 白いテキスト
            header_fill = PatternFill(start_color="1F4E78", fill_type="solid")  # 濃い青
            for cell in worksheet[1]:
                cell.font = bold_font
                cell.border = thin_border
                cell.fill = header_fill

            # ヘッダー以外には罫線を適用し、行ごとに色を交互に設定
            light_blue_fill = PatternFill(start_color="D9E1F2", fill_type="solid")  # 薄い青
            for row_num, row in enumerate(
                worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column),
                start=2,
            ):
                for cell in row:
                    cell.border = thin_border
                    # 偶数行に薄い青を適用（2,4,6...）
                    if row_num % 2 == 0:
                        cell.fill = light_blue_fill

            # 列の最大文字列長を計算して、文字折り返しが必要な列を判定
            max_char_limit = 20  # この値を超える列に対して文字折り返しを適用

            for col_num, column in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_num)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except TypeError:
                        pass

                # 文字列が長い列にのみ文字折り返しを適用
                if max_length > max_char_limit:
                    for cell in column:
                        cell.alignment = Alignment(wrap_text=True)
                    # 列幅を適切に設定
                    worksheet.column_dimensions[column_letter].width = 30
                else:
                    # 短い列は自動調整
                    worksheet.column_dimensions[column_letter].width = max_length + 2

            # ヘッダーの行を固定
            worksheet.freeze_panes = "A2"

            # アノテーション対象列以外は読み取り専用にする
            for col_num, column in enumerate(worksheet.columns, 1):
                # この列がアノテーション対象かを判定
                is_annotation_col = self.columns[col_num - 1] in self.annotation_cols

                for cell in column:
                    # ヘッダー行は常にロックし、アノテーション対象列のデータ行のみ編集可能にする
                    is_header = cell.row == 1
                    cell.protection = Protection(locked=is_header or not is_annotation_col)

            # シートを保護
            worksheet.protection.sheet = True
            worksheet.protection.enable()

            # Schemaが設定されている場合，アノテーション対象列に入力規制を追加する
            if hasattr(self, "schema") and self.schema is not None:
                for col_name in self.annotation_cols:
                    col = self.schema.get_column(col_name)
                    if col is not None and col.allowed_values is not None:
                        col_idx = list(self.columns).index(col_name) + 1
                        col_letter = get_column_letter(col_idx)
                        formula = ",".join(str(v) for v in col.allowed_values)
                        dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
                        dv.add(f"{col_letter}2:{col_letter}{worksheet.max_row}")
                        worksheet.add_data_validation(dv)


if __name__ == "__main__":  # テンプレートの作成例
    df = pd.DataFrame(
        {
            "id": [1, 2, 3],
            "text": ["This is a sample text.", "Another example.", "More data."],
            "label": [None, None, None],
        }
    )
    template = Template.from_dataframe(df, id_cols=["id"], annotation_cols=["label"])
    template.to_excel("output/annotation_template.xlsx")
