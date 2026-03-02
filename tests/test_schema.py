import pandas as pd
import pytest
from openpyxl import load_workbook

from excelano.annotation_data import AnnotationData
from excelano.schema import Column, Schema, SchemaValidationError
from excelano.template import Template


class TestColumn:
    """Columnクラスのテスト"""

    def test_column_validate_allowed_values_valid(self):
        """許容値のバリデーションが正常に通ることを確認"""
        col = Column(name="relevance", dtype=int, allowed_values=[0, 1])
        series = pd.Series([0, 1, 0, 1])
        errors = col.validate(series)
        assert errors == []

    def test_column_validate_allowed_values_invalid(self):
        """許容値以外の値がエラーになることを確認"""
        col = Column(name="relevance", dtype=int, allowed_values=[0, 1])
        series = pd.Series([0, 1, 2, 1])
        errors = col.validate(series)
        assert len(errors) == 1
        assert "3 行目" in errors[0]
        assert "'relevance'" in errors[0]
        assert "2" in errors[0]

    def test_column_validate_allowed_values_multiple_errors(self):
        """複数の不正値がすべて検出されることを確認"""
        col = Column(name="relevance", dtype=int, allowed_values=[0, 1])
        series = pd.Series([0, 5, 2, 1])
        errors = col.validate(series)
        assert len(errors) == 2
        assert "2 行目" in errors[0]
        assert "3 行目" in errors[1]

    def test_column_validate_skips_null_values(self):
        """Null値がスキップされることを確認"""
        col = Column(name="relevance", allowed_values=[0, 1])
        series = pd.Series([0, None, 1])
        errors = col.validate(series)
        assert errors == []

    def test_column_validate_dtype_valid(self):
        """dtype変換可能な値がバリデーションを通ることを確認"""
        col = Column(name="value", dtype=int)
        series = pd.Series([1, 2, 3])
        errors = col.validate(series)
        assert errors == []

    def test_column_validate_dtype_invalid(self):
        """dtypeに変換できない値がエラーになることを確認"""
        col = Column(name="value", dtype=int)
        series = pd.Series(["abc", 2, 3])
        errors = col.validate(series)
        assert len(errors) == 1
        assert "1 行目" in errors[0]
        assert "'value'" in errors[0]
        assert "abc" in errors[0]

    def test_column_validate_no_constraints(self):
        """制約なしの場合にバリデーションが通ることを確認"""
        col = Column(name="text")
        series = pd.Series(["hello", "world", 123])
        errors = col.validate(series)
        assert errors == []

    def test_column_validate_string_dtype(self):
        """文字列型のdtypeバリデーションを確認"""
        col = Column(name="name", dtype=str)
        series = pd.Series(["Alice", "Bob", "Charlie"])
        errors = col.validate(series)
        assert errors == []


class TestSchema:
    """Schemaクラスのテスト"""

    def test_schema_validate_valid_data(self):
        """正常なデータがバリデーションを通ることを確認"""
        schema = Schema(
            columns=[
                Column(name="id", dtype=int),
                Column(name="relevance", dtype=int, allowed_values=[0, 1]),
            ],
            id_cols=["id"],
            annotation_cols=["relevance"],
        )
        df = pd.DataFrame({"id": [1, 2, 3], "relevance": [0, 1, 0]})
        errors = schema.validate(df)
        assert errors == []

    def test_schema_validate_duplicate_ids(self):
        """重複IDがエラーになることを確認"""
        schema = Schema(
            columns=[Column(name="id", dtype=int)],
            id_cols=["id"],
            annotation_cols=[],
        )
        df = pd.DataFrame({"id": [1, 1, 3]})
        errors = schema.validate(df)
        assert len(errors) == 1
        assert "一意に識別できません" in errors[0]

    def test_schema_validate_invalid_values(self):
        """不正な値がエラーとして検出されることを確認"""
        schema = Schema(
            columns=[
                Column(name="id", dtype=int),
                Column(name="relevance", dtype=int, allowed_values=[0, 1]),
            ],
            id_cols=["id"],
            annotation_cols=["relevance"],
        )
        df = pd.DataFrame({"id": [1, 2, 3], "relevance": [0, 1, 5]})
        errors = schema.validate(df)
        assert len(errors) == 1
        assert "3 行目" in errors[0]

    def test_schema_get_dtype_dict(self):
        """get_dtype_dictが正しい辞書を返すことを確認"""
        schema = Schema(
            columns=[
                Column(name="id", dtype=int),
                Column(name="text", dtype=str),
                Column(name="label"),
            ],
            id_cols=["id"],
            annotation_cols=["label"],
        )
        dtype_dict = schema.get_dtype_dict()
        assert dtype_dict == {"id": int, "text": str}

    def test_schema_get_column(self):
        """get_columnが正しいColumnオブジェクトを返すことを確認"""
        col = Column(name="relevance", dtype=int, allowed_values=[0, 1])
        schema = Schema(columns=[col], id_cols=["id"], annotation_cols=["relevance"])
        assert schema.get_column("relevance") is col
        assert schema.get_column("nonexistent") is None


class TestSchemaValidationError:
    """SchemaValidationErrorのテスト"""

    def test_error_contains_all_messages(self):
        """エラーメッセージがすべて含まれていることを確認"""
        errors = ["エラー1", "エラー2"]
        exc = SchemaValidationError(errors)
        assert exc.errors == errors
        assert "エラー1" in str(exc)
        assert "エラー2" in str(exc)


class TestAnnotationDataWithSchema:
    """AnnotationDataとSchema統合のテスト"""

    def test_from_excel_with_schema_valid(self, tmp_path):
        """Schemaバリデーションが通る正常なデータの読み込みを確認"""
        df = pd.DataFrame(
            {
                "id": [1, 2, 3],
                "text": ["a", "b", "c"],
                "relevance": [0, 1, 0],
            }
        )
        file_path = tmp_path / "data.xlsx"
        df.to_excel(file_path, index=False)

        schema = Schema(
            columns=[
                Column(name="id", dtype=int),
                Column(name="text", dtype=str),
                Column(name="relevance", dtype=int, allowed_values=[0, 1]),
            ],
            id_cols=["id"],
            annotation_cols=["relevance"],
        )

        result = AnnotationData.from_excel(
            file_path=str(file_path),
            dtype={"id": int, "text": str, "relevance": int},
            annotated_cols=["relevance"],
            id_cols=["id"],
            schema=schema,
        )
        assert isinstance(result, AnnotationData)
        assert result.schema is schema

    def test_from_excel_with_schema_invalid_values(self, tmp_path):
        """Schemaバリデーションで不正値がエラーになることを確認"""
        df = pd.DataFrame(
            {
                "id": [1, 2, 3],
                "text": ["a", "b", "c"],
                "relevance": [0, 1, 5],
            }
        )
        file_path = tmp_path / "data.xlsx"
        df.to_excel(file_path, index=False)

        schema = Schema(
            columns=[
                Column(name="relevance", dtype=int, allowed_values=[0, 1]),
            ],
            id_cols=["id"],
            annotation_cols=["relevance"],
        )

        with pytest.raises(SchemaValidationError) as exc_info:
            AnnotationData.from_excel(
                file_path=str(file_path),
                dtype={"id": int, "text": str, "relevance": int},
                annotated_cols=["relevance"],
                id_cols=["id"],
                schema=schema,
            )
        assert "3 行目" in str(exc_info.value)

    def test_from_excel_without_schema_still_works(self, tmp_path):
        """Schema無しの既存APIが引き続き動作することを確認"""
        df = pd.DataFrame(
            {
                "id": [1, 2, 3],
                "text": ["a", "b", "c"],
                "relevance": [0, 1, 0],
            }
        )
        file_path = tmp_path / "data.xlsx"
        df.to_excel(file_path, index=False)

        result = AnnotationData.from_excel(
            file_path=str(file_path),
            dtype={"id": int, "text": str, "relevance": int},
            annotated_cols=["relevance"],
            id_cols=["id"],
        )
        assert isinstance(result, AnnotationData)


class TestTemplateWithSchema:
    """TemplateとSchema統合のテスト"""

    def test_from_dataframe_with_schema(self):
        """Schema付きでTemplateが作成できることを確認"""
        schema = Schema(
            columns=[
                Column(name="id", dtype=int),
                Column(name="label", allowed_values=[0, 1]),
            ],
            id_cols=["id"],
            annotation_cols=["label"],
        )
        df = pd.DataFrame({"id": [1, 2, 3], "text": ["a", "b", "c"], "label": [None, None, None]})
        template = Template.from_dataframe(df, id_cols=["id"], annotation_cols=["label"], schema=schema)
        assert template.schema is schema

    def test_from_csv_with_schema(self, tmp_path):
        """Schema付きでCSVからTemplateが作成できることを確認"""
        csv_file = tmp_path / "data.csv"
        df = pd.DataFrame({"id": [1, 2, 3], "text": ["a", "b", "c"], "label": [None, None, None]})
        df.to_csv(str(csv_file), index=False)

        schema = Schema(
            columns=[Column(name="id", dtype=int)],
            id_cols=["id"],
            annotation_cols=["label"],
        )
        template = Template.from_csv(str(csv_file), id_cols=["id"], annotation_cols=["label"], schema=schema)
        assert template.schema is schema

    def test_to_excel_with_data_validation(self, tmp_path):
        """to_excelでアノテーション列に入力規制が設定されることを確認"""
        schema = Schema(
            columns=[
                Column(name="id", dtype=int),
                Column(name="label", allowed_values=[0, 1]),
            ],
            id_cols=["id"],
            annotation_cols=["label"],
        )
        df = pd.DataFrame({"id": [1, 2, 3], "text": ["a", "b", "c"], "label": [None, None, None]})
        template = Template.from_dataframe(df, id_cols=["id"], annotation_cols=["label"], schema=schema)

        file_path = tmp_path / "template.xlsx"
        template.to_excel(str(file_path))

        workbook = load_workbook(str(file_path))
        worksheet = workbook.active

        # データバリデーションが設定されていることを確認
        assert len(worksheet.data_validations.dataValidation) == 1
        dv = worksheet.data_validations.dataValidation[0]
        assert dv.type == "list"
        assert dv.formula1 == '"0,1"'

    def test_to_excel_without_schema_no_data_validation(self, tmp_path):
        """Schema無しの場合にデータバリデーションが設定されないことを確認"""
        df = pd.DataFrame({"id": [1, 2, 3], "text": ["a", "b", "c"], "label": [None, None, None]})
        template = Template.from_dataframe(df, id_cols=["id"], annotation_cols=["label"])

        file_path = tmp_path / "template.xlsx"
        template.to_excel(str(file_path))

        workbook = load_workbook(str(file_path))
        worksheet = workbook.active

        # データバリデーションが設定されていないことを確認
        assert len(worksheet.data_validations.dataValidation) == 0

    def test_from_dataframe_with_schema_validation_error(self):
        """SchemaバリデーションエラーがTemplate作成時に発生することを確認"""
        schema = Schema(
            columns=[
                Column(name="text", dtype=str, allowed_values=["a", "b", "c"]),
            ],
            id_cols=["id"],
            annotation_cols=["label"],
        )
        df = pd.DataFrame({"id": [1, 2, 3], "text": ["a", "b", "invalid"], "label": [None, None, None]})

        with pytest.raises(SchemaValidationError) as exc_info:
            Template.from_dataframe(df, id_cols=["id"], annotation_cols=["label"], schema=schema)
        assert "3 行目" in str(exc_info.value)
