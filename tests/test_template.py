import pandas as pd
import pytest
from openpyxl import load_workbook

from excelano.template import Template


@pytest.fixture
def sample_template():
    """テスト用のサンプルTemplateを作成"""
    df = pd.DataFrame(
        {
            "id": [1, 2, 3],
            "text": ["Short", "Medium length text", "This is a very long text that should trigger wrapping"],
            "label": [None, None, None],
        }
    )
    return Template.from_dataframe(df, id_cols=["id"], annotation_cols=["label"])


def test_to_excel_creates_file(sample_template, tmp_path):
    """to_excelメソッドがファイルを作成することを確認"""
    file_path = tmp_path / "test_output.xlsx"
    sample_template.to_excel(str(file_path))
    assert file_path.exists()


def test_to_excel_header_styling(sample_template, tmp_path):
    """ヘッダーのスタイルが正しく適用されていることを確認"""
    file_path = tmp_path / "test_output.xlsx"
    sample_template.to_excel(str(file_path))

    workbook = load_workbook(str(file_path))
    worksheet = workbook.active

    # ヘッダーセル（1行目）を確認
    header_cell = worksheet["A1"]

    # フォントが太字で白色であることを確認
    assert header_cell.font.bold is True
    assert header_cell.font.color.rgb == "00FFFFFF"  # FFFFFF = 白

    # 背景色が濃い青であることを確認
    assert header_cell.fill.start_color.rgb == "001F4E78"  # 1F4E78 = 濃い青

    # 罫線が適用されていることを確認
    assert header_cell.border.left.style == "thin"
    assert header_cell.border.right.style == "thin"
    assert header_cell.border.top.style == "thin"
    assert header_cell.border.bottom.style == "thin"


def test_to_excel_data_rows_border(sample_template, tmp_path):
    """データ行に罫線が適用されていることを確認"""
    file_path = tmp_path / "test_output.xlsx"
    sample_template.to_excel(str(file_path))

    workbook = load_workbook(str(file_path))
    worksheet = workbook.active

    # データ行のセル（2行目以降）を確認
    data_cell = worksheet["A2"]

    assert data_cell.border.left.style == "thin"
    assert data_cell.border.right.style == "thin"
    assert data_cell.border.top.style == "thin"
    assert data_cell.border.bottom.style == "thin"


def test_to_excel_alternating_row_colors(sample_template, tmp_path):
    """データ行に交互に色が適用されていることを確認"""
    file_path = tmp_path / "test_output.xlsx"
    sample_template.to_excel(str(file_path))

    workbook = load_workbook(str(file_path))
    worksheet = workbook.active

    # 偶数行（2,4,6...）に薄い青が適用されていることを確認
    even_row_cell = worksheet["A2"]
    assert even_row_cell.fill.start_color.rgb == "00D9E1F2"  # D9E1F2 = 薄い青

    # 奇数行には背景色が適用されていないことを確認
    odd_row_cell = worksheet["A3"]
    assert odd_row_cell.fill.start_color.index == 0 or odd_row_cell.fill.fill_type is None


def test_to_excel_annotation_col_protection(sample_template, tmp_path):
    """アノテーション対象列がロック解除されていることを確認"""
    file_path = tmp_path / "test_output.xlsx"
    sample_template.to_excel(str(file_path))

    workbook = load_workbook(str(file_path))
    worksheet = workbook.active

    # annotation_cols = ["label"] なので、label列（C列）はロック解除
    label_cell = worksheet["C2"]
    assert label_cell.protection.locked is False


def test_to_excel_non_annotation_col_protection(sample_template, tmp_path):
    """アノテーション対象外の列がロックされていることを確認"""
    file_path = tmp_path / "test_output.xlsx"
    sample_template.to_excel(str(file_path))

    workbook = load_workbook(str(file_path))
    worksheet = workbook.active

    # annotation_cols = ["label"] なので、id列（A列）とtext列（B列）はロック
    id_cell = worksheet["A2"]
    text_cell = worksheet["B2"]

    assert id_cell.protection.locked is True
    assert text_cell.protection.locked is True


def test_to_excel_sheet_protection_enabled(sample_template, tmp_path):
    """シート保護が有効になっていることを確認"""
    file_path = tmp_path / "test_output.xlsx"
    sample_template.to_excel(str(file_path))

    workbook = load_workbook(str(file_path))
    worksheet = workbook.active

    # シート保護が有効であることを確認
    assert worksheet.protection.sheet is True


def test_to_excel_freeze_panes(sample_template, tmp_path):
    """ヘッダー行が固定されていることを確認"""
    file_path = tmp_path / "test_output.xlsx"
    sample_template.to_excel(str(file_path))

    workbook = load_workbook(str(file_path))
    worksheet = workbook.active

    # freeze_panesが"A2"に設定されていることを確認
    assert worksheet.freeze_panes == "A2"


def test_to_excel_long_text_wrapping(sample_template, tmp_path):
    """長いテキストを含む列に文字折り返しが適用されることを確認"""
    file_path = tmp_path / "test_output.xlsx"
    sample_template.to_excel(str(file_path))

    workbook = load_workbook(str(file_path))
    worksheet = workbook.active

    # text列（B列）は長いテキストを含むため文字折り返しが適用
    text_cell = worksheet["B2"]
    assert text_cell.alignment.wrap_text is True


def test_to_excel_column_width(sample_template, tmp_path):
    """列幅が適切に設定されていることを確認"""
    file_path = tmp_path / "test_output.xlsx"
    sample_template.to_excel(str(file_path))

    workbook = load_workbook(str(file_path))
    worksheet = workbook.active

    # 長いテキストを含む列の幅がwrap_width + 2に設定されていることを確認
    assert worksheet.column_dimensions["B"].width == sample_template.wrap_width + 2


def test_from_dataframe_with_validation(tmp_path):
    """from_dataframeメソッドが検証ロジックを正しく実行することを確認"""
    # 重複するidを含むDataFrame
    df_duplicate = pd.DataFrame(
        {
            "id": [1, 1, 3],
            "text": ["text1", "text2", "text3"],
            "label": [None, None, None],
        }
    )

    with pytest.raises(ValueError, match="id_colsで指定された列の組み合わせで一意に識別できません"):
        Template.from_dataframe(df_duplicate, id_cols=["id"], annotation_cols=["label"])


def test_from_dataframe_with_annotation_col_values(tmp_path):
    """アノテーション対象列に値が入っている場合のエラーを確認"""
    df_with_values = pd.DataFrame(
        {
            "id": [1, 2, 3],
            "text": ["text1", "text2", "text3"],
            "label": [1, None, None],  # labelにすでに値が入っている
        }
    )

    with pytest.raises(ValueError, match="annotation_colsで指定された列に値が入っています"):
        Template.from_dataframe(df_with_values, id_cols=["id"], annotation_cols=["label"])


def test_from_csv(tmp_path):
    """from_csvメソッドが正しく動作することを確認"""
    csv_file = tmp_path / "test_data.csv"
    df = pd.DataFrame(
        {
            "id": [1, 2, 3],
            "text": ["text1", "text2", "text3"],
            "label": [None, None, None],
        }
    )
    df.to_csv(str(csv_file), index=False)

    template = Template.from_csv(str(csv_file), id_cols=["id"], annotation_cols=["label"])

    assert isinstance(template, Template)
    assert template.annotation_cols == ["label"]
    assert template.id_cols == ["id"]
    assert len(template) == 3
